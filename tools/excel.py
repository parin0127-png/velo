import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sqlite3
import os

DB_PATH = "velo.db"

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn
    
def header_style(cell , bg = "636cff"):
    cell.font = Font(bold = True, color = "ffffff", size = 11)
    cell.fill = PatternFill("solid" , fgColor = bg)
    cell.alignment = Alignment(horizontal = "center" , vertical = "center")

def data_style(cell):
    cell.font = Font(size = 10)
    cell.alignment = Alignment(vertical = "center")

def generate_excel():
    output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)) , ".." , "output")
    os.makedirs(output_folder , exist_ok = True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_folder , f"Velo_Report_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    conn = get_conn()
    cursor = conn.cursor()

    ws1 = wb.active
    ws1.title = "Overview"

    ws1.merge_cells("A1:B1")
    ws1["A1"].value = f"Velo Business Overview - {datetime.now().strftime('%d %B %Y')}"
    ws1["A1"].font = Font(bold = True, size = 13, color = "6c63ff")

    overview_header = ["Metric", "Value"]
    for i , h in enumerate(overview_header,1):
        header_style(ws1.cell(row = 2 , column = i, value = h))
    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = [t[0] for t in cursor.fetchall()]

    total_clients = 0
    active = 0
    follow_up = 0
    pending = 0
    total_tasks = 0

    if "clients" in tables:
        cursor.execute("SELECT COUNT(*) FROM clients"); total_clients = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM clients WHERE status='active'"); active = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM clients WHERE status='follow up'"); follow_up = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM clients WHERE status='pending'"); pending = cursor.fetchone()[0]

    if "tasks" in tables:
            cursor.execute("SELECT COUNT(*) FROM tasks"); total_tasks = cursor.fetchone()[0]

    overview_data = [
        ("Total Clients" , total_clients),
        ("Active Clients" , active),
        ("Follow up Requried" , follow_up),
        ("Pending Clients" , pending),
        ("Total Tasks Run" , total_tasks)
    ]

    for i , (metric , value) in enumerate(overview_data, 3):
        ws1.cell(row = i, column = 1, value = metric).font = Font(bold = True, size = 10)
        ws1.cell(row = i, column = 2, value = value).alignment = Alignment(horizontal = "center")
    
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 15

    ws2 = wb.create_sheet("Clients")

    ws2.merge_cells("A1:E1")
    ws2["A1"].value = f"Client Report {datetime.now().strftime('%d %B %Y')}"
    ws2["A1"].font = Font(bold = True , size = 13 , color = "6c63ff")

    client_headers = ["Name" , "Email" , "Status" , "Last contact" , "Notes"]
    for i , h in enumerate(client_headers, 1):
        header_style(ws2.cell(row = 2, column = i, value = h))
    
    if "clients" in tables:
        cursor.execute("SELECT name, email, status, last_contact, notes FROM clients")
        for row_idx, row in enumerate(cursor.fetchall(),3):
            for col_idx, value in enumerate(row, 1):
                cell = ws2.cell(row = row_idx , column = col_idx, value = value)
                data_style(cell)
                if col_idx == 3:
                    if value == "active":
                        cell.fill = PatternFill("solid" , fgColor = "d1fae5")
                    elif value == "follow up":
                        cell.fill = PatternFill("solid" , fgColor = "fef3c7")
                    elif value == "pending":
                        cell.fill = PatternFill("solid" , fgColor = "fee2e2")
    
    for col in ["A" , "B" , "C" , "D" , "E"]:
        ws2.column_dimensions[col].width = 22
    

    conn.close()
    wb.save(filepath)
    print(f"> Excel saved -> {filepath}")
    return filepath